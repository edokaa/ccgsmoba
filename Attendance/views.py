from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from .models import Mass, Member, This_Sunday_Member, Town
from django.urls import reverse
import datetime
from .forms import MemberForm, LoginForm
import openpyxl
from os import path
from django.db.utils import IntegrityError

from django.views import View
from django.views.generic import ListView

from django.contrib.auth import authenticate, login
# from django.contrib.auth.models import User


from django.contrib.auth.mixins import LoginRequiredMixin


# Create your views here.


today = datetime.date.today().strftime('%A, %B %d, %Y')

class IndexView(View):
    def get(self, request):
        title = 'CCGS Mass Attendance'
        context = {
            'title': title,
        }

        return render(request, 'index.html', context)


class LoginView(View):
    def get(self, request):
        login_form = LoginForm
        title = 'Admin Login'
        context = {
            'login_form': login_form,
            'title': title
        }
        return render(request,'login.html', context)

    def post(self, request, *args, **kwargs):
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return HttpResponseRedirect(reverse('attendance:attendance'))
        else:
            login_msg = '[x] invalid login details'

        login_form = LoginForm
        title = 'Admin Login'
        context = {
            'login_form': login_form,
            'title': title,
            'login_msg': login_msg,
        }
        return render(request, 'login.html', context)


class AttendanceView(LoginRequiredMixin, ListView):
    login_url = '/attendance/login/'
    redirect_field_name = 'attendance/'

    template_name = 'attendance.html'

    def get_queryset(self):
        return Mass.objects.all()

    def get_context_data(self, **kwargs):
        title = 'CCGS Mass Attendance'
        context = super().get_context_data(**kwargs)
        context['title'] = title
        context['today'] = today
        return context


class MassView(LoginRequiredMixin, View):
    login_url = 'attendance/login/'
    redirect_field_name = 'attendance/'

    def get(self, request, pk, *args, **kwargs):
        return HttpResponseRedirect(reverse('attendance:attendance'))

    def post(self, request, pk, *args, **kwargs):
        # pk = request.POST['pk']
        mass_name = Mass.objects.get(id=pk)
        title = f"Welcome to {mass_name}"
        no_of_members = no_of_mem(mass_name)
        mass_members_set = This_Sunday_Member.objects.filter(mass_attended_id=pk)[:10]
        context = {
            'mass_name': mass_name,
            'mass_members_set': mass_members_set,
            'no_of_members': no_of_members,
            'title': title,
            'today': today,
            'pk': pk,
        }
        return render(request, 'mass.html', context)


class SelectMemberView(LoginRequiredMixin, View):
    login_url = 'attendance/login/'
    redirect_field_name = 'attendance/'

    def get(self, request, pk, *args, **kwargs):
        return HttpResponseRedirect(reverse('attendance:attendance'))

    def post(self, request, pk, *args, **kwargs):
        if 'select_member' in request.POST:
            mass_name = Mass.objects.get(id=pk)
            title = f"Add Members to {mass_name}"
            no_of_members = no_of_mem(mass_name)
            context = {
                'title': title,
                'mass_name': mass_name,
                'no_of_members': no_of_members,
                'today': today,
                'pk': pk,
            }
            if 'member_id' in request.POST:
                mass_name = Mass.objects.get(id=pk)
                title = f"Add Members to {mass_name}"
                selected_member = Member.objects.get(id=request.POST['member_id'])
                mass_name = Mass.objects.get(id=pk)
                this_mass = Mass.objects.get(mass_name=mass_name)
                check_query = This_Sunday_Member.objects.all().filter(
                    member_name_id=selected_member.id,
                    member_address_id=selected_member.member_address_id,
                    mass_attended_id=this_mass.id
                )
                try:
                    if check_query:
                        error_msg = f"{selected_member} already added for mass, add someone else"
                        context = {
                            'selected_member': selected_member,
                            'mass_name': mass_name,
                            'no_of_members': no_of_members,
                            'title': title,
                            'error_msg': error_msg,
                            'today': today,
                            'pk': pk,
                        }
                    else:
                        This_Sunday_Member.objects.create(
                            member_name_id=selected_member.id,
                            member_phone_no=selected_member.member_phone_no,
                            member_address_id=selected_member.member_address_id,
                            mass_attended_id=this_mass.id
                        )

                        context = {
                            'selected_member': selected_member,
                            'mass_name': mass_name,
                            'title': title,
                            'no_of_members': no_of_members+1,
                            'today': today,
                            'pk': pk,
                        }
                    return render(request, 'select_member.html', context)

                except IntegrityError:
                    error_msg = f"{selected_member} has already been added to the previous mass, add someone else"
                    context = {
                        'selected_member': selected_member,
                        'mass_name': mass_name,
                        'title': title,
                        'error_msg': error_msg,
                        'no_of_members': no_of_members,
                        'today': today,
                        'pk': pk,
                    }
                    return render(request, 'select_member.html', context)
            return render(request, 'select_member.html', context)

        elif '_name_' in request.POST:
            mass_name = Mass.objects.get(id=pk)
            no_of_members = no_of_mem(mass_name)
            search_name = request.POST['_name_']
            queryset = Member.objects.all().order_by('-member_name').filter(
                member_name__icontains=search_name,
            )
            context = {
                'queryset': queryset,
                'mass_name': mass_name,
                'search_name': search_name,
                'no_of_members': no_of_members,
                'today': today,
                'pk': pk,

                # 'form': form,
            }
            return render(request, 'select_member.html', context)


class AddNewView(LoginRequiredMixin, View):
    login_url = 'attendance/login/'
    redirect_field_name = 'attendance/'

    def get(self, request, *args, **kwargs):
        return HttpResponseRedirect(reverse('attendance:index'))

    def post(self, request, pk, *args, **kwargs):
        mass_name = Mass.objects.get(id=pk)
        title = 'Add a New Member'
        form = MemberForm
        no_of_members = no_of_mem(mass_name)
        context = {
            'form': form,
            'title': title,
            'mass_name': mass_name,
            'no_of_members': no_of_members,
            'today': today,
            'pk': pk,
        }
        if 'submit_new' in request.POST:
            # if form.is_valid():

            member_check = not Member.objects.all().filter(
                member_name=request.POST['member_name'],
                member_address_id=request.POST['member_address']
            )
            if member_check:
                add_new_mem = Member.objects.create(
                    member_name=request.POST['member_name'],
                    member_phone_no=request.POST['member_phone_no'],
                    member_address_id=request.POST['member_address'],
                )
                if add_new_mem:
                    success_msg = 'Member Successfully added to Database'
                    this_mass = Mass.objects.get(mass_name=mass_name)
                    selected_member = Member.objects.get(member_name=request.POST['member_name'])
                    This_Sunday_Member.objects.create(member_name_id=selected_member.id,
                                                      member_phone_no=selected_member.member_phone_no,
                                                      member_address_id=selected_member.member_address_id,
                                                      mass_attended_id=this_mass.id)
                    no_of_members = no_of_mem(mass_name)
                    context = {
                        'form': form,
                        'title': title,
                        'mass_name': mass_name,
                        'success_msg': success_msg,
                        'no_of_members': no_of_members,
                        'today': today,
                        'pk': pk,
                    }
                else:
                    success_msg = 'There was an error adding Member to Database'
                    context = {
                        'form': form,
                        'title': title,
                        'mass_name': mass_name,
                        'success_msg': success_msg,
                        'no_of_members': no_of_members,
                        'pk': pk,
                    }
            else:
                error_msg = 'Member already exists'
                context = {
                    'form': form,
                    'title': title,
                    'mass_name': mass_name,
                    'error_msg': error_msg,
                    'no_of_members': no_of_members,
                    'today': today,
                    'pk': pk,
                }
        return render(request, 'new_member.html', context)


class FinalView(LoginRequiredMixin, View):
    login_url = 'attendance/login/'
    redirect_field_name = 'attendance/'

    def get(self, request, *args, **kwargs):
        title = 'Todays Details'
        mass_list = Mass.objects.all()
        mass_no_of_mem = [1,2,3]
        # for i,mass in enumerate(mass_list):
        #     mass_no_of_mem[i] = no_of_mem(mass)
        i = 0
        while i < len(mass_list):
            mass_no_of_mem[i] = no_of_mem(mass_list[i])
            i += 1
        total_count = This_Sunday_Member.objects.count()
        context = {
            'mass_list': mass_list,
            'title': title,
            'mass_no_of_mem': mass_no_of_mem,
            'today': today,
            'total_count': total_count,

        }
        return render(request, 'final_analysis.html', context)

    def post(self, request, *args, **kwargs):
        mass_list = Mass.objects.all()
        wb_path = 'today_log.xlsx'
        wb = load_workbook(wb_path)

        sheet_title = datetime.date.today().strftime('%d-%m-%Y')

        # if sheet_title in wb.sheetnames:
        #     wb[sheet_title].title = sheet_title + 'old'
        # else:
        wb.create_sheet(title=sheet_title)

        sheet = wb[sheet_title]

        sheet.append(('CATHOLIC CHURCH OF THE GOOD SHEPHERD, MOBA',))
        sheet.append((f"{today} - MASS ATTENDANCE",))
        sheet.append(('',))
        sheet.append((f"TOTAL: {This_Sunday_Member.objects.count()} MEMBERS",))
        sheet.append(('',))

        for mass in mass_list:
            sheet.append((mass.mass_name,))
            mass_members_set = This_Sunday_Member.objects.filter(
                mass_attended_id=Mass.objects.get(mass_name=mass).id)

            for i, mem in enumerate(mass_members_set, start=1):
                name = mem.member_name
                phone = mem.member_phone_no
                address = mem.member_address
                member = (i, name.member_name, phone, address.town_name)
                sheet.append(member)

            sheet.append(('',))
            sheet.append(('',))

        try:
            wb.save(wb_path)
            This_Sunday_Member.objects.all().delete()

        except:
            return HttpResponse('unsuccessful')

        return HttpResponseRedirect(reverse('attendance:index'))

def add_member_from_excel():
    wb = openpyxl.load_workbook('today_log.xlsx')
    sheet = wb['Sheet1']
    for value in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=3, values_only=True):
        name = value[0]
        if value[1] == None:
            phone = '-'
        else:
            phone = value[1]
        address = value[2]
        address_id = Town.objects.all().get(town_name=address).id
        if Member.objects.all().filter(member_name=name, member_phone_no=phone, member_address_id=address_id):
            continue
        else:
            Member.objects.create(member_name=name, member_phone_no=phone, member_address_id=address_id)


def members_for_each_mass(mass_name):
    mass_members_set = This_Sunday_Member.objects.filter(mass_attended_id=Mass.objects.get(mass_name=mass_name).id)
    print(mass_name)
    for i, mem in enumerate(mass_members_set, start=1):
        print(f"({i}) {mem.member_name.member_name} - {mem.member_phone_no} - {mem.member_address.town_name}")
    print('')

def load_workbook(wb_path):
    if path.exists(wb_path):
        return openpyxl.load_workbook(wb_path)
    return openpyxl.Workbook()


def strip_space(word):
    y = 0
    stripped = ''
    while y < len(word):
        x = word[y]
        if x != ' ':
            stripped += x
        else:
            stripped += '-'
        y += 1
    return stripped


def no_of_mem(the_mass):
    no_of_members = 0
    this_mass = Mass.objects.get(mass_name=the_mass)
    mass_members_set = This_Sunday_Member.objects.filter(mass_attended_id=this_mass.id)
    for member in mass_members_set:
        no_of_members += 1
    return no_of_members

